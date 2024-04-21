import sys
import pandas as pd
from PyQt5.QtCore import QDir
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, \
    QListWidget, QFileDialog
from openpyxl import load_workbook
import subprocess


class DialogWindow(QWidget):
    """
    Диалоговое окно PyQt5 для создания сводки измерений из файла Excel.

    Атрибуты:
        path (str): Путь выбранного файла Excel.
        label (QLabel): Метка, отображающая "Путь:".
        line_edit (QLineEdit): Виджет строки ввода для отображения выбранного пути файла.
        button (QPushButton): Кнопка для открытия диалогового окна выбора файла.
        list_widget (QListWidget): Виджет для отображения списка листов в файле Excel.

    Методы:
        __init__: Инициализирует виджет DialogWindow.
        open_file_dialog: Открывает диалоговое окно для выбора файла Excel.
        populate_list_widget: Заполняет виджет списка именами листов из выбранного файла Excel.
        pivot_sheets: Обрабатывает выбранные листы из файла Excel и создает сводку, сохраняя ее в новом файле Excel.
    """

    def __init__(self):
        """
        Инициализирует виджет DialogWindow.

        Создает компоновку, виджеты и подключает сигналы к слотам.
        """
        super().__init__()
        self.setWindowTitle("Сводка по замерам")
        self.resize(400, 400)
        self.path = ""
        vertical_layout = QVBoxLayout(self)

        horizontal_layout1 = QHBoxLayout()
        vertical_layout.addLayout(horizontal_layout1)

        self.label = QLabel("Путь:")
        horizontal_layout1.addWidget(self.label)

        self.line_edit = QLineEdit()
        horizontal_layout1.addWidget(self.line_edit)

        self.button = QPushButton("Обзор...")
        self.button.clicked.connect(self.open_file_dialog)  # Подключение слота к событию нажатия
        horizontal_layout1.addWidget(self.button)

        vertical_layout2 = QVBoxLayout()
        vertical_layout.addLayout(vertical_layout2)

        label = QLabel("Список листов документа")
        vertical_layout2.addWidget(label)

        self.list_widget = QListWidget()
        vertical_layout2.addWidget(self.list_widget)

        execute_button = QPushButton("Выполнить")
        execute_button.clicked.connect(self.pivot_sheets)  # Подключение слота к событию нажатия
        vertical_layout2.addWidget(execute_button)

        self.setLayout(vertical_layout)

    def open_file_dialog(self):
        """
        Открывает диалоговое окно для выбора файла Excel.

        Обновляет атрибут path и заполняет виджет списка именами листов.
        """
        options = QFileDialog.Options()
        desktop_path = QDir().homePath()
        file_path, _ = QFileDialog.getOpenFileName(self, "Выберите файл Excel", desktop_path,
                                                   "Excel Files (*.xlsx *.xls)", options=options)
        if file_path:
            self.path = file_path
            self.line_edit.setText(file_path)
            self.populate_list_widget(file_path)

    def populate_list_widget(self, file_path):
        """
        Заполняет виджет списка именами листов из выбранного файла Excel.

        Аргументы:
            file_path (str): Путь выбранного файла Excel.
        """
        try:
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
            self.list_widget.clear()
            self.list_widget.addItems(sheet_names)
        except Exception as e:
            print(f"Произошла ошибка при загрузке файла Excel: {e}")

    def pivot_sheets(self):
        """
        Обрабатывает выбранные листы из файла Excel и создает сводку, сохраняя ее в новом файле Excel.

        Создает сводную таблицу из выбранных листов, сохраняет ее в новом файле Excel и открывает файл.
        """
        try:
            workbook = load_workbook(self.path)
            sheet_names = workbook.sheetnames
            dfs = {}
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]

                data = sheet.values
                for _ in range(4):
                    next(data)
                columns = next(data)

                df = pd.DataFrame(data, columns=columns)[-1:]
                df = df[df.columns[4:]]
                df.dropna(axis=1)
                text = sheet_name
                well, kust, deposit, start_data = text.split('_')
                df_row = pd.DataFrame({'Дата начала замера': [start_data],
                                       'Месторождение': [deposit],
                                       'Куст': [kust],
                                       'Скважина': [well]})
                for colname in df.columns:
                    df_row[colname] = df[colname].values
                dfs[sheet_name] = df_row
            workbook.close()

            result_df = pd.concat([dfs[name] for name in dfs.keys()])
            self.path = self.path[:-5] + '__свод.xlsx'
            result_df.to_excel(self.path, index=False)
            subprocess.Popen(['start', self.path], shell=True)

        except Exception as ex:
            print('Требуется закрыть путь', self.path)
            print(ex)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DialogWindow()
    window.show()
    sys.exit(app.exec_())
